"""
Allocator blueprint — upload, configure, run allocation, view results, download.
"""
import io
import json
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, session, send_file, jsonify, current_app,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..services.sheets_service import load_master_timetable, get_available_slots
from ..services.allocation_engine import allocate
from ..models import db, AllocationRun

allocator_bp = Blueprint("allocator", __name__)


# ── Step 1 & 2: Upload Google Sheet URL + Roll Numbers ──

@allocator_bp.route("/upload", methods=["GET", "POST"])
def upload():
    """Accept Google Sheet URL and roll numbers."""
    if request.method == "POST":
        sheet_url = request.form.get("sheet_url", "").strip()
        roll_text = request.form.get("roll_numbers", "").strip()

        if not sheet_url:
            flash("Please paste a Google Sheet URL.", "danger")
            return render_template("allocator/upload.html")

        if not roll_text:
            flash("Please enter at least one roll number.", "danger")
            return render_template("allocator/upload.html")

        # Parse roll numbers (comma, newline, or space separated)
        import re
        roll_numbers = [
            r.strip()
            for r in re.split(r"[,\n\r\s]+", roll_text)
            if r.strip()
        ]

        # Store in session for next step
        session["sheet_url"] = sheet_url
        session["roll_numbers"] = roll_numbers

        flash(f"Loaded {len(roll_numbers)} roll numbers.", "success")
        return redirect(url_for("allocator.configure"))

    return render_template("allocator/upload.html")


# ── Step 3: Configure allocation parameters ──

@allocator_bp.route("/configure", methods=["GET", "POST"])
def configure():
    """Configure day filter, panels, slot limit, and select slots."""
    sheet_url = session.get("sheet_url")
    if not sheet_url:
        flash("Please start by entering a Google Sheet URL.", "warning")
        return redirect(url_for("allocator.upload"))

    if request.method == "POST":
        day_filter = request.form.get("day_filter", "").strip()
        panel_count = int(request.form.get("panel_count", 1))
        slot_limit = int(request.form.get("slot_limit", 10))
        selected_slots = request.form.getlist("selected_slots")

        if panel_count < 1:
            flash("Panel count must be at least 1.", "danger")
            return redirect(url_for("allocator.configure"))
        if slot_limit < 1:
            flash("Slot limit must be at least 1.", "danger")
            return redirect(url_for("allocator.configure"))
        if not selected_slots:
            flash("Please select at least one time slot.", "danger")
            return redirect(url_for("allocator.configure"))

        session["day_filter"] = day_filter
        session["panel_count"] = panel_count
        session["slot_limit"] = slot_limit
        session["selected_slots"] = selected_slots

        return redirect(url_for("allocator.run_allocation"))

    # The frontend app.js will now make an AJAX call to /api/slots to fetch the data
    # without blocking the initial page load.
    day_filter = request.args.get("day_filter", "")

    return render_template(
        "allocator/configure.html",
        available_slots=[],
        day_filter=day_filter,
    )


# ── AJAX: Get slots for a given day filter ──

@allocator_bp.route("/api/slots", methods=["GET"])
def api_slots():
    """Return available slots as JSON for the given day filter."""
    sheet_url = session.get("sheet_url")
    day_filter = request.args.get("day_filter", "")

    if not sheet_url:
        return jsonify({"error": "No sheet URL in session"}), 400

    try:
        slots = get_available_slots(sheet_url, day_filter)
        return jsonify({"slots": slots})
    except KeyError as e:
        current_app.logger.error("Missing column in sheet", exc_info=e)
        return jsonify({"error": f"Sheet column error: {e}. Check that the sheet has columns: roll_no, day, slot."}), 400
    except ValueError as e:
        current_app.logger.error("Sheet value error", exc_info=e)
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        err_msg = str(e).lower()
        if "403" in err_msg or "permission" in err_msg:
            return jsonify({"error": "Permission denied. Share the Google Sheet with the service account email."}), 403
        current_app.logger.error("Error fetching slots via API", exc_info=e)
        return jsonify({"error": f"Error reading sheet: {e}"}), 500


# ── Step 4: Run the allocation engine ──

@allocator_bp.route("/run", methods=["GET"])
def run_allocation():
    """Execute the allocation engine and display results."""
    sheet_url = session.get("sheet_url")
    roll_numbers = session.get("roll_numbers", [])
    day_filter = session.get("day_filter", "")
    panel_count = session.get("panel_count", 1)
    slot_limit = session.get("slot_limit", 10)
    selected_slots = session.get("selected_slots", [])

    if not sheet_url or not roll_numbers:
        flash("Missing input data. Please start over.", "warning")
        return redirect(url_for("allocator.upload"))

    try:
        df_master = load_master_timetable(sheet_url)
        results = allocate(
            df_master=df_master,
            roll_numbers=roll_numbers,
            day_filter=day_filter,
            panel_count=panel_count,
            slot_limit=slot_limit,
            selected_slots=selected_slots,
        )
    except Exception as e:
        current_app.logger.error("Allocation engine failure", exc_info=e)
        flash(f"Allocation error: {e}", "danger")
        return redirect(url_for("allocator.configure"))

    # Store results in session and DB
    session["results"] = results

    run_record = AllocationRun(sheet_url=sheet_url)
    run_record.set_config({
        "day_filter": day_filter,
        "panel_count": panel_count,
        "slot_limit": slot_limit,
        "selected_slots": selected_slots,
        "roll_count": len(roll_numbers),
    })
    run_record.set_results(results)
    db.session.add(run_record)
    db.session.commit()

    return redirect(url_for("allocator.results"))


# ── Step 5: Display results dashboard ──

@allocator_bp.route("/results")
def results():
    """Display the allocation results in a 3-section dashboard."""
    results = session.get("results")
    if not results:
        flash("No results found. Please run an allocation first.", "warning")
        return redirect(url_for("allocator.upload"))

    panel_count = session.get("panel_count", 1)
    day_filter = session.get("day_filter", "")

    # Group allocated students by slot for display
    slots_data = {}
    for entry in results.get("allocated", []):
        slot = entry["slot"]
        if slot not in slots_data:
            slots_data[slot] = []
        slots_data[slot].append(entry)

    return render_template(
        "allocator/results.html",
        results=results,
        slots_data=slots_data,
        panel_count=panel_count,
        day_filter=day_filter,
    )


# ── Step 6: Export to Excel ──

@allocator_bp.route("/download")
def download():
    """Generate a styled .xlsx file and send it for download."""
    results = session.get("results")
    if not results:
        flash("No results to export.", "warning")
        return redirect(url_for("allocator.upload"))

    panel_count = session.get("panel_count", 1)
    wb = _generate_excel(results, panel_count)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Mock_GD-PI_Allocation.xlsx",
    )


# ── Dashboard / History ──

@allocator_bp.route("/dashboard")
def dashboard():
    """Show dashboard with recent allocation runs."""
    runs = (
        AllocationRun.query
        .order_by(AllocationRun.timestamp.desc())
        .limit(10)
        .all()
    )
    return render_template("dashboard.html", runs=runs)


# ── Excel generation helper ──

def _generate_excel(results: dict, panel_count: int) -> Workbook:
    """
    Generate a styled Excel workbook mirroring the VBA output format.
    Three sections: Allocated, Not Available, Overflow, plus Summary.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mock GD-PI Allocation"

    # ── Styles ──
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    slot_fill = PatternFill(start_color="C8DCFF", end_color="C8DCFF", fill_type="solid")
    sub_header_fill = PatternFill(start_color="DCE6F0", end_color="DCE6F0", fill_type="solid")
    red_font = Font(bold=True, color="FF0000", size=12)
    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    red_header_fill = PatternFill(start_color="F0C8C8", end_color="F0C8C8", fill_type="solid")
    orange_font = Font(bold=True, color="FF8C00", size=12)
    orange_fill = PatternFill(start_color="FFF5DC", end_color="FFF5DC", fill_type="solid")
    orange_header_fill = PatternFill(start_color="FFE6C8", end_color="FFE6C8", fill_type="solid")
    blue_font = Font(bold=True, color="0000FF", size=12)
    blue_fill = PatternFill(start_color="DCE6FF", end_color="DCE6FF", fill_type="solid")
    summary_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    center = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    row = 1

    # ── Title ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="STUDENT ALLOCATION TO MOCK GD / PI SLOTS").font = title_font
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # ── Allocated slots ──
    # Group by slot
    slots_data = {}
    for entry in results.get("allocated", []):
        slot = entry["slot"]
        slots_data.setdefault(slot, []).append(entry)

    for slot_name in sorted(slots_data.keys(), key=lambda s: _parse_slot_time_xl(s)):
        entries = slots_data[slot_name]

        # Slot header
        cols = 4 if panel_count > 1 else 3
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=f"Time Slot: {slot_name}")
        cell.font = header_font
        cell.fill = slot_fill
        cell.alignment = center
        row += 1

        # Column headers
        headers = ["Roll No", "Name", "Batch"]
        if panel_count > 1:
            headers.append("Panel")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        # Data rows
        for entry in entries:
            ws.cell(row=row, column=1, value=entry["roll_no"]).alignment = center
            ws.cell(row=row, column=2, value=entry["name"]).alignment = center
            ws.cell(row=row, column=3, value=entry["batch"]).alignment = center
            if panel_count > 1:
                ws.cell(row=row, column=4, value=f"Panel {entry['panel']}").alignment = center
            for ci in range(1, cols + 1):
                ws.cell(row=row, column=ci).border = thin_border
            row += 1

        # Total row
        ws.cell(row=row, column=1, value="Total Allocated:").font = header_font
        ws.cell(row=row, column=2, value=len(entries)).font = header_font
        for ci in range(1, cols + 1):
            ws.cell(row=row, column=ci).fill = summary_fill
            ws.cell(row=row, column=ci).border = thin_border
            ws.cell(row=row, column=ci).alignment = center
        row += 2

    # ── Not Available Section ──
    not_available = results.get("not_available", [])
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="STUDENTS NOT AVAILABLE FOR ALLOCATION")
    cell.font = red_font
    cell.fill = red_fill
    cell.alignment = center
    row += 2

    headers = ["Roll No", "Name", "Batch", "Reason"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = red_header_fill
        cell.alignment = center
        cell.border = thin_border
    row += 1

    if not_available:
        for entry in not_available:
            ws.cell(row=row, column=1, value=entry["roll_no"]).alignment = center
            ws.cell(row=row, column=2, value=entry["name"]).alignment = center
            ws.cell(row=row, column=3, value=entry["batch"]).alignment = center
            ws.cell(row=row, column=4, value=entry["reason"]).alignment = center
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = thin_border
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value="All students have free slots in selected times!")
        cell.font = Font(bold=True, color="008000")
        cell.alignment = center
        row += 1

    # ── Overflow Section ──
    overflow = results.get("overflow", [])
    if overflow:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value="STUDENTS WITH FREE SLOTS BUT NOT ALLOCATED")
        cell.font = orange_font
        cell.fill = orange_fill
        cell.alignment = center
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value="(Reason: Slot capacity limits reached)")
        cell.font = Font(italic=True)
        cell.alignment = center
        row += 1

        headers = ["Roll No", "Name", "Batch", "Eligible Free Slots"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = header_font
            cell.fill = orange_header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        for entry in overflow:
            ws.cell(row=row, column=1, value=entry["roll_no"]).alignment = center
            ws.cell(row=row, column=2, value=entry["name"]).alignment = center
            ws.cell(row=row, column=3, value=entry["batch"]).alignment = center
            ws.cell(row=row, column=4, value=entry.get("eligible_slots", "")).alignment = center
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).border = thin_border
            row += 1

    # ── Summary Section ──
    summary = results.get("summary", {})
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value="ALLOCATION SUMMARY")
    cell.font = blue_font
    cell.fill = blue_fill
    cell.alignment = center
    row += 1

    summary_items = [
        ("Total Students in RollList:", summary.get("total_roll", 0), None),
        ("Students Allocated:", summary.get("allocated_count", 0), Font(bold=True, color="008000")),
        ("Students Not Available:", summary.get("not_available_count", 0), Font(bold=True, color="FF0000")),
    ]
    if summary.get("overflow_count", 0) > 0:
        summary_items.append(
            ("Available but Not Allocated:", summary["overflow_count"], Font(bold=True, color="FFA500"))
        )

    for label, value, font in summary_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.alignment = center
        if font:
            cell.font = font
        for ci in (1, 2):
            ws.cell(row=row, column=ci).border = thin_border
        row += 1

    # Auto-fit columns
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 25

    return wb


def _parse_slot_time_xl(slot: str) -> float:
    """Parse slot time for sorting."""
    try:
        start_str = slot.split("-")[0].strip()
        parts = start_str.replace(".", ":").split(":")
        return int(parts[0]) + int(parts[1]) / 60.0 if len(parts) > 1 else int(parts[0])
    except (ValueError, IndexError):
        return 9999.0
